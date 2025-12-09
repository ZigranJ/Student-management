import openpyxl
import os
import math

filePath = input("Enter file path: ").strip()

if not os.path.exists(filePath):
    print("File not found.")
    exit()

try:
    wb = openpyxl.load_workbook(filePath)
except PermissionError:
    print("Cannot open file. It may be open in Excel.")
    exit()

sheet = wb.active

while True:
    cmd = input("> ").strip()

    if cmd == "average":
        average_col = None

        for cell in sheet[1]:
            if isinstance(cell.value, str) and cell.value.lower() == "average":
                average_col = cell.column_letter

        if not average_col:
            average_col = openpyxl.utils.get_column_letter(sheet.max_column + 1)
            sheet[average_col + "1"] = "Average"

        start_col = 2
        end_col = openpyxl.utils.column_index_from_string(average_col) - 1

        for row in range(2, sheet.max_row + 1):
            values = []
            for col in range(start_col, end_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)):
                    values.append(cell_value)

            if values:
                sheet[average_col + str(row)] = sum(values) / len(values)

    elif cmd == "save":
        try:
            wb.save(filePath)
        except PermissionError:
            print("Cannot save file. It is probably open in Excel.")
    elif cmd == "top":
        average_col = None

        for cell in sheet[1]:
            if isinstance(cell.value, str) and cell.value.lower() == "average":
                average_col = cell.column_letter

        if not average_col:
            average_col = openpyxl.utils.get_column_letter(sheet.max_column + 1)
            sheet[average_col + "1"] = "Average"

        start_col = 2
        end_col = openpyxl.utils.column_index_from_string(average_col) - 1

        maxAverageFound = -math.inf
        foundRow = -1

        for row in range(2, sheet.max_row + 1):
            values = []
            for col in range(start_col, end_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)):
                    values.append(cell_value)

            if values:
                average = sum(values) / len(values)
                sheet[average_col + str(row)] = average

                if average > maxAverageFound:
                    maxAverageFound = average
                    foundRow = row

        print(sheet.cell(row=foundRow, column=1).value)

    elif cmd == "sort":
        average_col = None

        for cell in sheet[1]:
            if isinstance(cell.value, str) and cell.value.lower() == "average":
                average_col = cell.column_letter

        if not average_col:
            average_col = openpyxl.utils.get_column_letter(sheet.max_column + 1)
            sheet[average_col + "1"] = "Average"

        start_col = 2
        end_col = openpyxl.utils.column_index_from_string(average_col) - 1

        for row in range(2, sheet.max_row + 1):
            values = []
            for col in range(start_col, end_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)):
                    values.append(cell_value)

            if values:
                average = sum(values) / len(values)
                sheet[average_col + str(row)] = average
        
        avg_col_index = openpyxl.utils.column_index_from_string(average_col)

        data = []
        for row in range(2, sheet.max_row + 1):
            row_values = [sheet.cell(row=row, column=col).value
                          for col in range(1, sheet.max_column + 1)]
            data.append(row_values)

        data.sort(key=lambda r: (r[avg_col_index - 1] is not None, r[avg_col_index - 1]), reverse=True)

        for i, row_values in enumerate(data, start=2):
            for col, value in enumerate(row_values, start=1):
                sheet.cell(row=i, column=col).value = value

    elif cmd == "add":
        name = input("Name: ")

        lastRow = sheet.max_row + 1
        sheet.cell(row=lastRow, column=1).value = name

        for col in range(2, sheet.max_column + 1):
            sheet.cell(row=lastRow, column=col).value = 0

    elif cmd == "exit":
        break

try:
    wb.save(filePath)
except PermissionError:
    print("Cannot save file. It is probably open in Excel.")
